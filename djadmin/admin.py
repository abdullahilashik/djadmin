from django.contrib import admin
from .models import JobCategory, Job
from django.http import HttpResponse
import csv
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from openpyxl import load_workbook


class CSVCategoryForm(forms.Form):
    file = forms.FileField()


@admin.register(Job)
class JobAdmin(admin.ModelAdmin):
    pass


@admin.register(JobCategory)
class JobCategoryAdmin(admin.ModelAdmin):
    list_display = ('title', 'job_count')
    actions = ['delete_category', 'export_as_csv']
    change_list_template = 'admin/change_list.html'

    def delete_category(self, request, queryset):
        queryset.delete()
        return queryset

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset

    def job_count(self, obj):
        return obj.job_set.count()

    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        fieldnames = [field.name for field in meta.fields]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename={}.csv'.format(meta)
        writer = csv.writer(response)
        self.message_user(request, 'Dataset has been exported!')

        return response

    def import_data(self, requests):
        if requests.method == 'POST':
            file = requests.FILES['file']
            excel = load_workbook(filename=file.file)
            sheet = excel.active

            for item in sheet.iter_rows(min_row=2):
                data = [i.value for i in item]
                JobCategory.objects.update_or_create(
                    id=data[0],
                    title=data[1],
                    slug=data[2]
                )
            self.message_user(requests, 'Dataset has been imported!')

        form = CSVCategoryForm()
        context = {
            'form': form
        }
        return render(requests, 'admin/category-import.html', context)

    def get_urls(self):
        urls = super().get_urls()
        myurls = [
            path('import-csv/', self.import_data),
        ]
        return myurls + urls

    export_as_csv.short_description = 'Export Selected'
    delete_category.short_description = 'Delete Categories'
